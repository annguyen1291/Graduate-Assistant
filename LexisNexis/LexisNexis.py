from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from difflib import SequenceMatcher
import re
import time


fileName = "D:/Users/hans/Desktop/Graduate Assistant/LexisNexis.xlsx"
wb = load_workbook(fileName)
ws = wb.active
# driver = webdriver.PhantomJS("C:/Users/hans/Google Drive/Desktop/Graduate Assistant/phantomjs-2.1.1-windows/bin/phantomjs.exe")# driver = webdriver.PhantomJS("C:/Users/hans/Google Drive/Desktop/Graduate Assistant/phantomjs-2.1.1-windows/bin/phantomjs.exe")
driver = webdriver.Chrome("D:/Users/hans/Desktop/Graduate Assistant/chromedriver.exe")
wait = WebDriverWait(driver, 15)
translate = "translate(text(), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ')"


def startup():
    driver.get("http://www-lexisnexis-com.ezproxy.baylor.edu/hottopics/lnacademic/?")
    driver.find_element_by_xpath("//input[@id='username']").send_keys("an_nguyen2")
    driver.find_element_by_xpath("//input[@id='password']").send_keys("9148807Aa")
    driver.find_element_by_xpath("//button[@type='submit'][text()='Log In']").send_keys(Keys.RETURN)
    xPath = "//a[text()='LexisNexis® Academic']"
    waitForXpath(xPath)


def search(row):
    driver.get("http://www-lexisnexis-com.ezproxy.baylor.edu/hottopics/lnacademic/?verb=sf&sfi=AC02NBCmpDosSrch")
    switchToFrame1()
    searchField = driver.find_element(By.XPATH, "//input[@id='companyname']")
    searchField.clear()
    global companyName
    companyName = str(ws['A' + str(row)].value)
    companyName = trimName(companyName)
    searchField.send_keys(companyName)
    searchField.send_keys(Keys.RETURN)
    try:
        driver.find_element(By.XPATH, "//ul[contains(text(), 'No companies found.')]").is_displayed()
        print("No companies found.")
        ws['L' + str(row)] = "n/a"
        ws['F' + str(row)] = "n/a"
        ws['G' + str(row)] = "n/a"
    except NoSuchElementException:
        try:
            driver.find_element(By.XPATH, "//ul/b[contains(text(), 'This Service is currently unavailable.')]").is_displayed()
            print("This Service is currently unavailable.")
            ws['L' + str(row)] = "n/a"
            ws['F' + str(row)] = "n/a"
            ws['G' + str(row)] = "n/a"
        except NoSuchElementException:
            try:
                driver.find_element(By.XPATH, "//ul[contains(text(), 'Your search has retrieved over 100,000 companies.')]").is_displayed()
                print("Your search has retrieved over 100,000 companies.")
                ws['L' + str(row)] = "n/a"
                ws['F' + str(row)] = "n/a"
                ws['G' + str(row)] = "n/a"
            except NoSuchElementException:
                try:
                    driver.find_element(By.XPATH, "//div[contains(text(), 'View All')]").click()
                except NoSuchElementException:
                    pass
                juxtapose(row)


def juxtapose(row):
    try:
        driver.find_element(By.XPATH, "//td[contains(text(), 'Find a Company - Results')]").is_displayed()
        print("multiple results")
        if ws['B' + str(row)].value is not None and ws['D' + str(row)].value is not None:
            print("Location available")
            locationAvailable1(row)
        else:
            print("Location not available")
            locationNotAvailable1(row)
    except NoSuchElementException:
        print("single results")
        switchToFrame2()
        if ws['B' + str(row)].value is not None and ws['D' + str(row)].value is not None:
            print("Location available")
            locationAvailable2(row)
        else:
            print("Location not available")
            locationNotAvailable2(row)


def locationAvailable1(row):
    try:
        location1('D', row).is_displayed()
        print("State matched")
        ws['N' + str(row)] = "matched"
        nameWeb = name1('D', row)
        nameMatch = compare(nameWeb.text.upper(), companyName)
        try:
            location1('B', row).is_displayed()
            print("City matched")
            ws['O' + str(row)] = "matched"
            if nameMatch < 0.8:
                ws['L' + str(row)] = nameWeb.text.upper()
                ws['M' + str(row)] = nameMatch
            nameWeb.send_keys(Keys.RETURN)
            switchToFrame2()
            info(row)
        except NoSuchElementException:
            print("City not matched")
            ws['O' + str(row)] = "n/m"
            compareName1(names1('D', row), 0)
    except NoSuchElementException:
        print("State not matched")
        ws['N' + str(row)] = "n/m"
        try:
            location1('B', row).is_displayed()
            print("City matched")
            ws['O' + str(row)] = "matched"
            nameWeb = name1('B', row)
            nameMatch = compare(nameWeb.text.upper(), companyName)
            if nameMatch > 0.8:
                nameWeb.send_keys(Keys.RETURN)
                switchToFrame2()
                info(row)
            else:
                ws['L' + str(row)] = "n/m"
        except NoSuchElementException:
            print("City not matched")
            ws['O' + str(row)] = "n/m"


def locationNotAvailable1(row):
    ws['N' + str(row)] = "n/a"
    ws['O' + str(row)] = "n/a"
    compareName1(driver.find_elements(By.XPATH, "//table[@class='resultList']//a"), 0)


def locationAvailable2(row):    
    a = False
    b = False
    try:
        a = location2('D', row).is_displayed()
        print("State matched")
        ws['N' + str(row)] = "matched"
    except NoSuchElementException:
        print("State not matched")
        ws['N' + str(row)] = "n/m"
    try:
        b = location2('B', row).is_displayed()
        print("City matched")
        ws['O' + str(row)] = "matched"
    except NoSuchElementException:
        print("City not matched")
        ws['O' + str(row)] = "n/m"
    if a + b >= 1:
        compareName2(name2().text.upper(), companyName)
    else:
        ws['L' + str(row)] = "n/a"


def locationNotAvailable2(row):
    ws['N' + str(row)] = "n/a"
    ws['O' + str(row)] = "n/a"
    compareName2(name2().text.upper(), companyName)


def info(row):
    time.sleep(2)
    sic = driver.find_element(By.XPATH, "//td[contains(text(), 'Primary SIC Code')]/parent::tr/td[2]/a")
    ws['F' + str(row)] = str(re.findall("\d+", sic.text))
    naics = driver.find_element(By.XPATH, "//td[contains(text(), 'Primary NAICS Code')]/parent::tr/td[2]")
    ws['G' + str(row)] = str(re.findall("\d+", naics.text))
    address = driver.find_element(By.XPATH, "//td[text()='Address:']/parent::tr/td[3]//tr[2]/td")
    ws['H' + str(row)] = str(re.findall("\d+", address.text))[:7] + "']"
    tryExcept("//td[text()='Type:']/parent::tr/td[3]", 'I', row)
    tryExcept("//td[text()='Employees:']/parent::tr/td[3]", 'J', row)
    tryExcept("//td[text()='Net Sales (USD)']/parent::tr/td[2]", 'K', row)
    wb.save(fileName)


def tryExcept(xpath, column, row):
    try:
        ws[column + str(row)] = driver.find_element(By.XPATH, xpath).text
    except NoSuchElementException:
        ws[column + str(row)] = "n/a"


def compare(object1, object2):
    return SequenceMatcher(None, object1, object2).ratio()


def location1(column, row):
    return driver.find_element(By.XPATH, "//table[@class='resultList']/tbody/tr/td[2][contains(" + translate + ", '" + ws[column + str(row)].value + "')]")


def name1(column, row):
    return driver.find_element(By.XPATH, "//table[@class='resultList']/tbody/tr/td[2][contains(" + translate + ", '" + ws[column + str(row)].value + "')]/parent::tr/td[1]//a")


def names1(column, row):
    return driver.find_elements(By.XPATH, "//table[@class='resultList']/tbody/tr/td[2][contains(" + translate + ", '" + ws[column + str(row)].value + "')]/parent::tr/td[1]//a")


def location2(column, row):
    return driver.find_element(By.XPATH, "//td[text()='Address:']/parent::tr/td[3]//tr[2]/td[contains(" + translate + ", '" + ws[column + str(row)].value + "')]")


def name2():
    return driver.find_element(By.XPATH, "//html/body/table/tbody/tr/td[1]/table[1]/tbody/tr[1]/td/table/tbody/tr/td/a")


def compareName1(nameWebs, nameMatch):
    count = len(nameWebs)
    if count > 4:
        count = 4
    time.sleep(count)
    for i in range(0, count):
        nameweb = trimName(nameWebs[i].text.upper())
        webMatch = compare(nameweb, companyName)
        if webMatch > nameMatch:
            nameMatch = webMatch
            nameTest = nameWebs[i]
    if nameMatch > 0.8:
        print("Name matched")
        ws['L' + str(row)] = "matched"
    else:
        print("Name not matched")
        ws['L' + str(row)] = nameTest.text.upper()
        ws['M' + str(row)] = "{0:.0f}%".format(nameMatch * 100)
    nameTest.send_keys(Keys.RETURN)
    switchToFrame2()
    info(row)


def compareName2(nameWeb, companyName):
    nameMatch = compare(trimName(nameWeb), companyName)
    if nameMatch < 0.8:
        ws['L' + str(row)] = nameWeb
        ws['M' + str(row)] = "{0:.0f}%".format(nameMatch * 100)
    info(row)


def trimName(companyName):
    companyName = companyName.replace(',', '')
    companyName = companyName.replace('.', '')
    last = len(companyName)
    surfix1 = ['(', 'DBA', 'D/B/A']
    for i in range(0, len(surfix1)):
        if companyName.find(surfix1[i]) > 0:
            last = companyName.find(surfix1[i])
    surfix2 = [" LLC", " INC", " LTD", " LLP"]
    for i in range(0, len(surfix2)):
        if companyName.find(surfix2[i]) > 0:
            last = companyName.find(surfix2[i])
    return companyName[:last]


def waitForXpath(inputXpath):
    wait = WebDriverWait(driver, 15)
    wait.until(EC.presence_of_element_located((By.XPATH, inputXpath)))


def switchToFrame1():
    wait.until(EC.frame_to_be_available_and_switch_to_it(driver.find_element(By.XPATH, "//iframe[@id='mainFrame']")))
    wait.until(EC.frame_to_be_available_and_switch_to_it(driver.find_element(By.XPATH, "//iframe[@id='dsrFrame']")))
    wait.until(EC.frame_to_be_available_and_switch_to_it(driver.find_element(By.XPATH, "//frame[@name='APPLICATION_CONTENT']")))


def switchToFrame2():
    wait.until(EC.frame_to_be_available_and_switch_to_it(driver.find_element(By.XPATH, "//iframe[@id='DOSSIER_CONTENT']")))
    wait.until(EC.frame_to_be_available_and_switch_to_it(driver.find_element(By.XPATH, "//frame[@name='CONTENT']")))
    wait.until(EC.frame_to_be_available_and_switch_to_it(driver.find_element(By.XPATH, "//frame[@name='BROWSE_DISPLAY']")))


startup()
for row in range(724, 882):
    #if ws['N' + str(row)].value == "matched" and ws['O' + str(row)].value == "n/m":
    print("\nrow: " + str(row))
    search(row)
    wb.save(fileName)
    time.sleep(1)
