# This script is for download the stock price of nearly 700 companies on Yahoo Finance
# Please remember to chagne the absolute paths to your own
from selenium import webdriver
from openpyxl import load_workbook
import time
import os
import csv

# Load excel file with the list of the companies which need searching
fileName = "D:/Users/hans/Desktop/Graduate Assistant/fs list 2017_An.xlsx"
wb = load_workbook(fileName)
ws = wb.active

# Initiate chrome web driver for web automation
driver = webdriver.Chrome("D:/Users/hans/Desktop/Graduate Assistant/chromedriver.exe")
driver.get("https://finance.yahoo.com/")
time.sleep(5)

# Extract the cache of the section at Yahoo Finance
driver.get("https://finance.yahoo.com/quote/AET/history?period1=1483164000&period2=1509426000&interval=1d&filter=history&frequency=1d")
link = driver.find_element_by_xpath("//a[span[text()='Download Data']]").get_attribute('href')
cache = link[(link.find("crumb") + 6):]

# Download a particular company's stock price by its stock code (row)
def run(row):
    print("\nrow: " + str(row))
    driver.get("https://query1.finance.yahoo.com/v7/finance/download/" + ws['D' + str(row)].value + "?period1=1483164000&period2=1509426000&interval=1d&events=history&crumb=" + cache)

# Loop through the list in the excel file
for row in range(2, 150):
    run(row)

# Merge all the downloaded excel into one file
directory = os.fsencode("C:/Users/hans/Desktop/Graduate Assistant/Yahoo Finance/")
YahooFianance = "C:/Users/hans/Desktop/Graduate Assistant/Yahoo Finance.csv"

with open(YahooFianance, 'a') as fout:
    writer = csv.writer(fout, delimiter=',')
    for file in os.listdir(directory):
    	# File name = stock code
        filename = os.fsdecode(file)
        with open("C:/Users/hans/Desktop/Graduate Assistant/Yahoo Finance/" + filename, 'r') as fin:
            reader = csv.reader(fin, delimiter=',')
            for line in reader:
            	# Add each line of individual files to the final one together with stock code
                writer.writerow([filename, line])
