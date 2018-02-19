from openpyxl import load_workbook
from difflib import SequenceMatcher


fileName = "D:/Users/hans/Desktop/Graduate Assistant/assignee and ultimate parent_name check_An 10212017.xlsx"
wb = load_workbook(fileName)
ws = wb.active


for row in range(2, 693):
	# print("\nrow: " + str(row))
	name1 = ws['K' + str(row)].value
	# print(name1)
	for i in name1.split():
		# print(i)
		loc = ws['O' + str(row)].value.find(i)
		# print(loc)
		if loc != -1:
			ws['Q' + str(row)] = 1
	name2 = ws['O' + str(row)].value
	# print(name2)
	for i in name2.split():
		# print(i)
		loc = ws['K' + str(row)].value.find(i)
		# print(loc)
		if loc != -1:
			ws['Q' + str(row)] = 1
	ws['P' + str(row)] = SequenceMatcher(None, name1, name2).ratio()	

wb.save(fileName)