from openpyxl import load_workbook

fileName = "C:/Users/hans/Google Drive/Desktop/Graduate Assistant/list - full.xlsx"
wb = load_workbook(fileName)
ws = wb.active

print(ws['M120'].value is not None)