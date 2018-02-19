from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from difflib import SequenceMatcher
import re

fileName = "C:/Users/hans/Google Drive/Desktop/Graduate Assistant/Hoover.xlsx"
wb = load_workbook(fileName)
ws = wb.active

# driver = webdriver.PhantomJS("C:/Users/hans/Google Drive/Desktop/Graduate Assistant/phantomjs-2.1.1-windows/bin/phantomjs.exe")
driver = webdriver.Chrome("C:/Users/hans/Google Drive/Desktop/Graduate Assistant/chromedriver.exe")
driver.get("http://research.library.gsu.edu/DB_HOOVER")
driver.find_element_by_xpath("//font[text()='CampusID']/parent::b/parent::td/parent::tr//input").send_keys("qwu4")
driver.find_element_by_xpath("//font[text()='Password']/parent::b/parent::td/parent::tr//input").send_keys("Nguyen_An2017")
driver.find_element_by_xpath("//input[@class='btnSubmit']").click()
driver.find_element_by_xpath("//a").click()


def search(row):
    try:
        driver.find_element_by_xpath("//div[contains(text(), 'All Categories')]").click()
        driver.find_element_by_xpath("//a[contains(text(), 'Companies')]").click()
    except NoSuchElementException:
        pass
    searchField = driver.find_element_by_xpath("//input[@id='searchField']")
    searchField.clear()
    global companyName
    companyName = str(ws['A' + str(row)].value)
    last = len(companyName)
    surfix1 = ['(', 'DBA', 'D/B/A']
    for i in range(0, len(surfix1)):
        if companyName.find(surfix1[i]) > 0:
            last = companyName.find(surfix1[i])
    surfix2 = [" LLC", " INC", " L.L.C.", " LTD", " L.P.", " LLP"]
    for i in range(0, len(surfix2)):
        if companyName.find(surfix2[i]) > 0:
            last = companyName.find(surfix2[i]) + len(surfix2[i])
    companyName = companyName[:last]
    searchField.send_keys(companyName)
    driver.find_element_by_xpath("//button[@id='btnSearch']").click()
    try:
        if driver.find_element_by_xpath("//body[@id='noResults']").is_displayed():
            print("No Results")
            ws['AC' + str(row)] = "n/a"
    except NoSuchElementException:
        wait("//tbody[@id='simpleSearchResults']")
        choose(row)


def choose(row):
    if ws['N' + str(row)].value is not None and ws['M' + str(row)].value is not None:
        print("Location available")
        locationAvailable(row)
    else:
        print("Location not available")
        locationNotAvailable(row)


def locationAvailable(row):
    try:
        stateWeb = element('N', row)
        print("State matched")
        ws['AE' + str(row)] = "matched"
        nameWeb = stateWeb.text.upper()
        nameMatch = compare(nameWeb, companyName)
        try:
            cityWeb = element('M', row)
            print("City matched")
            ws['AF' + str(row)] = "matched"
            if nameMatch < 0.8:
                ws['AC' + str(row)] = nameWeb
                ws['AD' + str(row)] = nameMatch
            stateWeb.click()
            info(row)
        except NoSuchElementException:
            print("City not matched")
            location = driver.find_element_by_xpath("//th[@id='companyName']/ancestor::table/tbody/tr/td[2][contains(text(), '" + ws['N' + str(row)].value + "')]").text
            cityWeb = location[:location.find(ws['N' + str(row)].value) - 1]
            cityMatch = compare(cityWeb, ws['M' + str(row)].value)
            ws['AF' + str(row)] = cityWeb
            ws['AG' + str(row)] = cityMatch
            if nameMatch > 0.8:
                stateWeb.click()
                info(row)
            else:
                ws['AC' + str(row)] = "n/m"
    except NoSuchElementException:
        print("State not matched")
        ws['AE' + str(row)] = "n/m"
        try:
            cityWeb = element('M', row)
            print("City matched")
            ws['AF' + str(row)] = "matched"
            nameWeb = cityWeb.text.upper()
            nameMatch = compare(nameWeb, companyName)
            if nameMatch > 0.8:
                ws['AC' + str(row)] = nameWeb
                ws['AD' + str(row)] = nameMatch
                cityWeb.click()
                info(row)
            else:
                ws['AC' + str(row)] = "n/m"
        except NoSuchElementException:
            print("City not matched")
            ws['AF' + str(row)] = "n/m"


def locationNotAvailable(row):
    ws['AE' + str(row)] = "n/a"
    ws['AF' + str(row)] = "n/a"
    results = driver.find_elements_by_xpath("//tbody/tr/td/a")
    nameMatch = 0
    for i in range(0, len(results)):
        webMatch = compare(results[i].text.upper(), companyName)
        if webMatch > nameMatch:
            nameMatch = webMatch
            nameTest = results[i]
    if nameMatch > 0.8:
        print("Name matched")
        ws['AC' + str(row)] = nameTest.text.upper()
        ws['AD' + str(row)] = nameMatch
        nameTest.click()
        info(row)
    else:
        print("Name not matched")
        ws['AC' + str(row)] = "n/m"


def info(row):
    duns = driver.find_element_by_xpath("//th[text()=' Number']/parent::tr/td")
    ws['P' + str(row)] = str(re.findall("\d+", duns.text))
    sic = driver.find_element_by_xpath("//th[text()='Primary SIC Code']/parent::tr/td")
    ws['Q' + str(row)] = str(re.findall("\d+", sic.text))
    naics = driver.find_element_by_xpath("//th[text()='Primary NAICS Code']/parent::tr/td")
    ws['R' + str(row)] = str(re.findall("\d+", naics.text))
    tryExcept("//th[text()='Year of Founding']/parent::tr/td", 'S', row)
    tryExcept("//dfn[contains(text(), 'Sales') and contains(text(), '(Estimated)')]/parent::th/parent::tr/td", 'T', row)
    tryExcept("//dfn[contains(text(), 'Sales') and contains(text(), '(Actual)')]/parent::th/parent::tr/td", 'U', row)
    tryExcept("//th[text()='Employees (All Sites)']/parent::tr/td", 'V', row)
    tryExcept("//th[text()='Employees (This Site)']/parent::tr/td", 'W', row)
    tryExcept("//th[contains(text(), 'Plant/Facility Size')]/parent::tr/td", 'X', row)
    tryExcept("//th[text()='State of Incorporation']/parent::tr/td", 'Y', row)
    tryExcept("//span[@class='zip']", 'Z', row)
    tryExcept("//th[contains(text(), 'Ultimate Parent D-U-N-S')]/parent::tr/td", 'AA', row)
    tryExcept("//th[contains(text(), 'Immediate Parent D-U-N-S')]/parent::tr/td", 'AB', row)
    wb.save(fileName)


def tryExcept(xpath, column, row):
    try:
        ws[column + str(row)] = driver.find_element_by_xpath(xpath).text
    except NoSuchElementException:
        ws[column + str(row)] = "n/a"


def compare(object1, object2):
    return SequenceMatcher(None, object1, object2).ratio()


def element(column, row):
    return driver.find_element_by_xpath("//th[@id='companyName']/ancestor::table/tbody/tr/td[2][contains(text(), '" + ws[column + str(row)].value + "')]/parent::tr/td/a")


def wait(inputXpath):
    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.XPATH, inputXpath)))


for row in range(3774, 4429):
    print("\nrow: " + str(row))
    search(row)
