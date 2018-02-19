from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook

fileName = "C:/Users/hans/Google Drive/Desktop/Graduate Assistant/list - full.xlsx"
wb = load_workbook(fileName)
ws = wb.active

driver = webdriver.Chrome("C:/Users/hans/chromedriver.exe")
driver.get("http://research.library.gsu.edu/DB_HOOVER")
driver.find_element_by_xpath("//font[text()='CampusID']/parent::b/parent::td/parent::tr//input").send_keys("qwu4")
driver.find_element_by_xpath("//font[text()='Password']/parent::b/parent::td/parent::tr//input").send_keys("Nguyen_An2017")
driver.find_element_by_xpath("//input[@class='btnSubmit']").click()
driver.find_element_by_xpath("//a").click()


def search(row):
    searchField = driver.find_element_by_xpath("//input[@id='searchField']")
    searchField.clear()
    searchField.send_keys(str(ws['P' + str(row)].value))
    driver.find_element_by_xpath("//button[@id='btnSearch']").click()
    driver.find_element_by_xpath("//tbody/tr/td/a").click()
    try:
        ws['T' + str(row)] = driver.find_element_by_xpath("//dfn[contains(text(), 'Sales') and contains(text(), '(Estimated)')]/parent::th/parent::tr/td").text
    except NoSuchElementException:
        ws['T' + str(row)] = "n/a"
    wb.save(fileName)


for row in range(2530, 4428):
    if ws['P' + str(row)].value is not None:
        print("\nrow: " + str(row))
        search(row)
