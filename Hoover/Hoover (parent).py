from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
import re
import time

fileName = "D:/Users/hans/Desktop/Graduate Assistant/Ultimate company_Qinxi_20171018.xlsx"
wb = load_workbook(fileName)
ws = wb.active

driver = webdriver.Chrome("D:/Users/hans/Desktop/Graduate Assistant/chromedriver.exe")
driver.get("http://research.library.gsu.edu/DB_HOOVER")
driver.find_element_by_xpath("//font[text()='CampusID']/parent::b/parent::td/parent::tr//input").send_keys("qwu4")
driver.find_element_by_xpath("//font[text()='Password']/parent::b/parent::td/parent::tr//input").send_keys("godblessNK1999!")
driver.find_element_by_xpath("//input[@class='btnSubmit']").click()
driver.find_element_by_xpath("//a").click()


def search(row):
    searchField = driver.find_element_by_xpath("//input[@id='searchField']")
    searchField.clear()
    searchField.send_keys(str(ws['A' + str(row)].value))
    driver.find_element_by_xpath("//button[@id='btnSearch']").click()

    name = driver.find_element_by_xpath("//tbody/tr/td/a")
    ws['B' + str(row)] = name.text
    # name.click()

    # sic = driver.find_element_by_xpath("//th[text()='Primary SIC Code']/parent::tr/td")
    # ws['I' + str(row)] = str(re.findall("\d+", sic.text))
    # naics = driver.find_element_by_xpath("//th[text()='Primary NAICS Code']/parent::tr/td")
    # ws['J' + str(row)] = str(re.findall("\d+", naics.text))
    # tryExcept("//th[text()='Year of Founding']/parent::tr/td", 'G', row)
    # tryExcept("//dfn[contains(text(), 'Sales') and contains(text(), '(Estimated)')]/parent::th/parent::tr/td", 'K', row)
    # tryExcept("//dfn[contains(text(), 'Sales') and contains(text(), '(Actual)')]/parent::th/parent::tr/td", 'L', row)
    # tryExcept("//th[contains(text(), 'Employees') and contains(text(), '(All Sites)')]/parent::tr/td", 'E', row)
    # tryExcept("//th[contains(text(), 'Employees') and contains(text(), '(This Site)')]/parent::tr/td", 'F', row)
    # tryExcept("//th[contains(text(), 'Plant/Facility Size')]/parent::tr/td", 'D', row)
    # tryExcept("//th[text()='State of Incorporation']/parent::tr/td", 'H', row)
    # tryExcept("//span[@class='zip']", 'M', row)
    time.sleep(5)
    wb.save(fileName)


def tryExcept(xpath, column, row):
    try:
        ws[column + str(row)] = driver.find_element_by_xpath(xpath).text
    except NoSuchElementException:
        ws[column + str(row)] = "n/a"


for row in range(88, 93):
    print("\nrow: " + str(row))
    search(row)
