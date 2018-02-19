# This script is for finding the founding date of over 3400 companies on Bloomberg
# Please remember to chagne the absolute paths to your own
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, NoSuchFrameException, StaleElementReferenceException
from openpyxl import load_workbook
from difflib import SequenceMatcher
import time

# Load excel file with the list of the companies which need searching
fileName = "D:/Users/hans/Desktop/Graduate Assistant/Bloomberg/public fintech sample_yr founding.xlsx"
wb = load_workbook(fileName)
ws = wb.active
driver = webdriver.Chrome("D:/Users/hans/Desktop/Graduate Assistant/Misc/chromedriver.exe")


# Search for a particular company by its name (row)
def search(row):
    global companyName
    companyName = str(ws['A' + str(row)].value).replace(" ", "+") + "+"
    companyCity = str(ws['G' + str(row)].value).replace(" ", "+") + "+"
    companyState = str(ws['H' + str(row)].value).replace(" ", "+") + "+"
    searchQuery = "site:bloomberg.com/research/stocks/private/snapshot.asp"

    driver.get("https://www.google.com/search?q=" + companyName + searchQuery)  # + companyCity + companyState
    # Wait to enter captcha by Google
    try:
        driver.switch_to.frame(0)
        recaptcha_anchor = driver.find_element_by_xpath("//span[@id='recaptcha-anchor']")
        recaptcha_anchor.click()
        print("recaptcha")
        try:
            while recaptcha_anchor.is_displayed():
                time.sleep(1)
        except StaleElementReferenceException:
            pass
        # WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//td[@id='tdSearchTab_Company']")))
    except NoSuchFrameException:
        pass
    # try:
    #     driver.find_element_by_xpath("//div[@id='recaptcha']").click()
    #     print("recaptcha")
    #     time.sleep(30)
    # except NoSuchElementException:
    #     pass

    # If there is a result from Google
    try:
        # Click the link and store it
        try:
            driver.find_element_by_xpath("(//div[@class='srg']//a)[1]").click()
        except NoSuchElementException:
            driver.find_element_by_xpath("(//div[@class='g']//a)[1]").click()
        ws['O' + str(row)] = driver.current_url
        # Get the founding date
        try:
            resultFoundingDate = driver.find_element_by_xpath("//span[@itemprop='foundingDate']").text
            ws['K' + str(row)] = resultFoundingDate
        except NoSuchElementException:
            pass
        # Check name matching
        resultName = driver.find_element_by_xpath("//span[@itemprop='name']").text.upper()
        ws['P' + str(row)] = resultName
        ws['Q' + str(row)] = SequenceMatcher(None, companyName, resultName).ratio()
        # Check location matching
        location(row)
    # If there is no result, mark the excel file
    except NoSuchElementException:
        print("not found")
        pass
        # ws['G' + str(row)] = "n/a"
        # ws['H' + str(row)] = "n/a"


# Check location matching
def location(row):
    ws['L' + str(row)] = driver.find_element_by_xpath("//div[@itemprop='address']").text
    location = driver.find_element_by_xpath("//div[@itemprop='address']/p[last() - 1]").text
    seperator = location.find(",")
    ws['M' + str(row)] = location[:seperator]
    ws['N' + str(row)] = location[seperator + 2:location.find(" ", seperator + 2)]
    # city = ws['B' + str(row)].value
    # state = ws['C' + str(row)].value
    # # If location is available for checking
    # if city is not None and state is not None:
    #     print("Location available")
    #     resultAddress = driver.find_element_by_xpath("//div[@itemprop='address']").text.upper()
    #     # Check city name in the address
    #     if city in resultAddress:
    #         ws['J' + str(row)] = "matched"
    #     else:
    #         ws['J' + str(row)] = resultAddress
    #     # Check state name in the address
    #     if state in resultAddress:
    #         ws['K' + str(row)] = "matched"
    #     else:
    #         ws['K' + str(row)] = resultAddress
    # # If location is not available for checking, mark the excel file
    # else:
    #     print("Location not available")
    #     ws['J' + str(row)] = "n/a"
    #     ws['K' + str(row)] = "n/a"


# Loop through the list in the excel file
for row in [2, 3, 4, 5, 7, 8, 9, 11, 13, 14, 16, 17, 19, 24, 26, 27, 36, 37, 42, 43, 46, 47, 61, 216, 217, 218, 219, 220, 221, 222, 223, 224, 225, 226, 227, 228, 229, 230, 231, 232
]:  # range(3, 233):
    print("\nrow: " + str(row))
    search(row)
    wb.save(fileName)
