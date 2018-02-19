# This script is for finding the founding date of over 3400 companies on Bloomberg
# Please remember to chagne the absolute paths to your own
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from difflib import SequenceMatcher
import time

# Load excel file with the list of the companies which need searching
fileName = "D:/Users/hans/Desktop/Graduate Assistant/Bloomberg/public fintech sample_yr founding.xlsx"
wb = load_workbook(fileName)
ws = wb.active
driver = webdriver.Chrome("D:/Users/hans/Desktop/Graduate Assistant/Misc/chromedriver.exe")


# Search for a particular company by its name (row)
def search(row):
    global companyName
    companyName = str(ws['A' + str(row)].value)
    driver.get("https://www.google.com/search?q=site:bloomberg.com/research/stocks/private/snapshot.asp+" + companyName.replace(" ", "+"))
    # Wait to enter captcha by Google
    try:
        driver.find_element_by_xpath("//span[@id='ctl00_lblFeedback']").is_displayed()
        print("recaptcha")
        while driver.find_element_by_xpath("//span[@id='ctl00_lblFeedback']").is_displayed():
            time.sleep(1)
        # WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//td[@id='tdSearchTab_Company']")))
    except NoSuchElementException:
        pass
    # try:
    #     driver.find_element_by_xpath("//div[@id='recaptcha']").click()
    #     print("recaptcha")
    #     time.sleep(30)
    # except NoSuchElementException:
    #     pass

    # If there is a result from Google
    try:
        # Click the link and store it
        driver.find_element_by_xpath("(//cite[contains(text(), 'research/stocks/private/snapshot')]//ancestor::div[@class='rc']//a)[1]").click()
        ws['L' + str(row)] = driver.current_url
        # Get the founding date
        resultFoundingDate = driver.find_element_by_xpath("//span[@itemprop='foundingDate']").text
        ws['G' + str(row)] = resultFoundingDate
        # Check name matching
        resultName = driver.find_element_by_xpath("//span[@itemprop='name']").text.upper()
        ws['H' + str(row)] = resultName
        ws['I' + str(row)] = SequenceMatcher(None, companyName, resultName).ratio()
        # Check location matching
        location(row)
    # If there is no result, mark the excel file
    except NoSuchElementException:
        print("not found")
        ws['G' + str(row)] = "n/a"
        ws['H' + str(row)] = "n/a"


# Check location matching
def location(row):
    city = ws['B' + str(row)].value
    state = ws['C' + str(row)].value
    # If location is available for checking
    if city is not None and state is not None:
        print("Location available")
        resultAddress = driver.find_element_by_xpath("//div[@itemprop='address']").text.upper()
        # Check city name in the address
        if city in resultAddress:
            ws['J' + str(row)] = "matched"
        else:
            ws['J' + str(row)] = resultAddress
        # Check state name in the address
        if state in resultAddress:
            ws['K' + str(row)] = "matched"
        else:
            ws['K' + str(row)] = resultAddress
    # If location is not available for checking, mark the excel file
    else:
        print("Location not available")
        ws['J' + str(row)] = "n/a"
        ws['K' + str(row)] = "n/a"


# Loop through the list in the excel file
for row in range(3352, 3408):
    print("\nrow: " + str(row))
    search(row)
    wb.save(fileName)
