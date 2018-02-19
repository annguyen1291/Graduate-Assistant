from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from difflib import SequenceMatcher
import time

# Load excel file with the list of the companies which need searching
fileName = "D:/Users/hans/Desktop/Graduate Assistant/Bizapedia/public fintech extra12052017.xlsx"
wb = load_workbook(fileName)
ws = wb.active
# chromeOptions = webdriver.ChromeOptions()
# chromeOptions.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
driver = webdriver.Chrome("D:/Users/hans/Desktop/Graduate Assistant/Misc/chromedriver.exe")  # , chrome_options=chromeOptions)


# Search for a particular company by its name (row)
def search(row):
    companyName = str(ws['B' + str(row)].value)
    driver.get("https://www.bizapedia.com/")

    captcha()

    city = ws['C' + str(row)].value
    state = ws['E' + str(row)].value
    if city is not None and state is not None:
        print("Location available")
        companyCity = city
        companyState = state
    else:
        companyCity = ""
        companyState = "Select State/Province"
        print("Location not available")
        # ws['J' + str(row)] = "n/a"
        # ws['K' + str(row)] = "n/a"

    enterName = driver.find_element_by_xpath("//input[@id='txtCompanyName_Company']")
    enterCity = driver.find_element_by_xpath("//input[@id='txtCity_Company']")
    selectState = driver.find_element_by_xpath("//select[@id='selStateProvince_Company']")
    enterName.send_keys(companyName)
    enterCity.send_keys(companyCity)
    Select(selectState).select_by_visible_text(companyState)
    enterName.send_keys(Keys.ENTER)
    # try:
    #     driver.find_element_by_xpath("//table[@id='tblSearchForm_Company']/tbody/tr/td[7]").click()
    # except:
    #     driver.find_element_by_xpath("//table[@id='tblSearchForm_Company']/tbody/tr/td[7]").send_keys(Keys.ENTER)

    captcha()

    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//td[@id='tdResults']/table[2]/tbody/tr/td[2]")))
        resultName = ""
        resultRatio = 0
        companyNames = driver.find_elements_by_xpath("//td[@id='tdResults']/table[2]/tbody/tr/td[2]")
        for name in companyNames:
            ratio = SequenceMatcher(None, fileName.text.upper(), companyName).ratio()
            if ratio > resultRatio:
                resultName = name.text
                resultRatio = ratio

        ws['K' + str(row)] = resultName
        ws['L' + str(row)] = resultRatio
        ws['M' + str(row)] = driver.find_element_by_xpath("//td[contains(text(), '" + resultName + "')]/parent::tr/td[4]").text
        ws['N' + str(row)] = driver.find_element_by_xpath("//td[contains(text(), '" + resultName + "')]/parent::tr/td[7]").text

        driver.find_element_by_xpath("//td[contains(text(), '" + resultName + "')]").click()

        captcha()

        ws['O' + str(row)] = driver.current_url
        try:
            ws['F' + str(row)] = driver.find_element_by_xpath("//td[contains(text(), 'File Number:')]/parent::tr/td[2]").text
        except NoSuchElementException:
            pass
        try:
            ws['G' + str(row)] = driver.find_element_by_xpath("//td[contains(text(), 'Filing State:')]/parent::tr/td[2]").text
        except NoSuchElementException:
            pass
        try:
            ws['H' + str(row)] = driver.find_element_by_xpath("//td[contains(text(), 'Filing Status:')]/parent::tr/td[2]").text
        except NoSuchElementException:
            pass
        try:
            ws['I' + str(row)] = driver.find_element_by_xpath("//td[contains(text(), 'Filing Date:')]/parent::tr/td[2]").text
        except NoSuchElementException:
            pass
        try:
            ws['J' + str(row)] = driver.find_element_by_xpath("//td[contains(text(), 'Company Age:')]/parent::tr/td[2]").text
        except NoSuchElementException:
            pass
    except TimeoutException:
        pass


def captcha():
    try:
        driver.find_element_by_xpath("//span[@id='ctl00_lblFeedback']").is_displayed()
        print("recaptcha")
        while driver.find_element_by_xpath("//span[@id='ctl00_lblFeedback']").is_displayed():
            time.sleep(1)
        # WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//td[@id='tdSearchTab_Company']")))
    except NoSuchElementException:
        pass


# Loop through the list in the excel file
for row in range(2, 21):  # [5, 6, 9, 20, 22, 27, 28, 29, 33, 34, 36, 37, 38, 39, 42, 46, 49, 54, 57, 59, 63, 67, 71, 74, 86, 88, 90, 96, 97, 100, 113, 118, 119, 120, 121, 127, 134, 135, 141, 144, 150, 151, 156, 165, 170, 171, 173, 174, 176, 177, 181, 183, 184, 188, 189, 193, 194, 215, 216]:#
    print("\nrow: " + str(row))
    search(row)
    wb.save(fileName)
